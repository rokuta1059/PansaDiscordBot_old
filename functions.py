#!/usr/bin/env python
# -*- coding: utf-8 -*-
import random
import openpyxl
import json
from bs4 import BeautifulSoup
import urllib.request
from datetime import datetime, date, time

def makeConch(conch):
    f = open('magicConch.txt', 'r', encoding="utf8")
    readlist = f.read()
    f.close()
    conch = readlist.split('\n')
    return conch

def absurbDiary(inputInt):
    jsonFile = open('diary.json', encoding='utf8').read()
    data = json.loads(jsonFile)
    if inputInt == 0:
        return data[str(random.randint(1, len(data)))]
    else:
        return data[str(inputInt)]

def absurbFindName(name):
    findName = []
    jsonFile = open('diary.json', encoding='utf8').read()
    data = json.loads(jsonFile)
    for i in range(1, len(data) + 1):
        if name in data[str(i)]['name']:
            findName.append(data[str(i)]['number'])
    return findName

def absurbFindAbsurb(search):
    findAb = []
    jsonFile = open('diary.json', encoding='utf8').read()
    data = json.loads(jsonFile)
    for i in range(1, len(data) + 1):
        if search in data[str(i)]['absurb'] or search in data[str(i)]['description']:
            findAb.append(data[str(i)]['number'])
    return findAb

def addAbsurb(absurb, name, descript):
    jsonFile = open('diary.json', encoding='UTF8').read()
    data = json.loads(jsonFile)
    tmp = {'absurb':absurb, 'name':name, 'description':descript, 'number':len(data)+1}
    data[str(len(data)+1)] = tmp

    with open('diary.json', 'w', encoding='UTF8') as saveFile:
        json.dump(data, saveFile, indent='\t', ensure_ascii=False)

    return tmp

def changeAbsurb(absurb, name, descript, num):
    jsonFile = open('diary.json', encoding='UTF8').read()
    data = json.loads(jsonFile)
    tmp = {'absurb':absurb, 'name':name, 'description':descript, 'number':num}
    data[str(num)] = tmp

    with open('diary.json', 'w', encoding='UTF8') as saveFile:
        json.dump(data, saveFile, indent='\t', ensure_ascii=False)

    return tmp

def cypersRank(nickname, game):
    parseNick = urllib.parse.quote(nickname)
    if game == "공식":
        cypersURL = 'http://cyphers.nexon.com/cyphers/game/record/search/1/' + parseNick + '/1'
    else:
        cypersURL = 'http://cyphers.nexon.com/cyphers/game/record/search/2/' + parseNick + '/2'
    with urllib.request.urlopen(cypersURL) as response:
        html = response.read()
        soup = BeautifulSoup(html, 'html.parser')

    a = []
    battle = []
    img = []

    for firstTable in soup.find_all("div", "info info1"):
        for firstValue in firstTable.find_all("td"):
            a.append(firstValue.get_text())

    for secondTable in soup.find_all("div", "info info2"):
        for secondValue in secondTable.find_all("td"):
            a.append(secondValue.get_text())

    for battleTable in soup.find_all("li", "show"):
        for battleResult in battleTable.find_all("td"):
            battle.append(battleResult.get_text().replace(
                '\n', '').replace('\r', '').replace('\t', ''))

        for battleResult in battleTable.find_all("td", "char"):
            for imgList in battleResult.find_all("img"):
                img.append(imgList.get("src"))

    return cypersURL, a, battle, img

def getMillDate(name):
    f = open('militaryDate.txt', 'r', encoding="utf8")
    mill = []
    findName = False
    while True:
        line = f.readline()
        if line.find(name) != -1:
            mill = line.replace('\n', '').split(' ')
            findName = True
            break
        if not line:
            break
    f.close()

    if findName:
        milldate = list(map(int, mill[1].split('.')))
        now = datetime.now()
        findDate = datetime(milldate[0], milldate[1], milldate[2]) - now
        return "{0}쿤의 전역일은 **{1}일** 남았답니다~".format(mill[0], findDate.days)
    else:
        return "{0}쿤은 군인이 아닌거 같아요~".format(name)

def todayWeather(search):
    client_id = 'NAVER_CLIENT_ID'
    client_secret = 'NAVER_CLIENT_SECRET'
    encText = urllib.parse.quote(search + ' 날씨')
    url = 'https://search.naver.com/search.naver?sm=top_hty&fbm=1&ie=utf8&query=' + encText
    request = urllib.request.Request(url)
    request.add_header("X-Naver-Client-Id",client_id)
    request.add_header("X-Naver-Client-Secret",client_secret)
    response = urllib.request.urlopen(request)
    bsObj = BeautifulSoup(response, "html.parser")
    todayBase = bsObj.find('div', {'class': 'today_area _mainTabContent'})

    temp = bsObj.find('div', {'class': 'select_box'})
    temp2 = temp.find('em')
    local = temp2.text.strip()

    temp = todayBase.find('span', {'class': 'todaytemp'})
    todayTemp = temp.text.strip()

    temp = todayBase.find('p', {'class': 'cast_txt'})
    todayCast = temp.text.strip()

    temp = todayBase.find('span', {'class': 'merge'})
    todayMerge = temp.text.strip()

    temp = todayBase.find('span', {'class': 'sensible'})
    temp2 = temp.find('span', {'class': 'num'})
    todaySensible = temp2.text.strip()

    temp = todayBase.find('dl', {'class': 'indicator'})
    temp2 = temp.find_all('dd')
    todayDust = temp2[0].text.strip()

    weatherTable = [local, todayTemp, todayCast, todayMerge, todaySensible, todayDust]

    return weatherTable

def tomorrowWeather(search):
    client_id = 'NAVER_CLIENT_ID'
    client_secret = 'NAVER_CLIENT_SECRET'
    encText = urllib.parse.quote(search + ' 날씨')
    url = 'https://search.naver.com/search.naver?sm=top_hty&fbm=1&ie=utf8&query=' + encText
    request = urllib.request.Request(url)
    request.add_header("X-Naver-Client-Id",client_id)
    request.add_header("X-Naver-Client-Secret",client_secret)
    response = urllib.request.urlopen(request)
    bsObj = BeautifulSoup(response, "html.parser")

    temp = bsObj.find('div', {'class': 'select_box'})
    temp2 = temp.find('em')
    local = temp2.text.strip()

    tomorrowBase = bsObj.find_all('div', {'class': 'main_info morning_box'})

    temp = tomorrowBase[0].find('span', {'class': 'todaytemp'})
    tomorrowTemp = temp.text.strip()

    temp = tomorrowBase[0].find('ul', {'class': 'info_list'})
    temp2 = temp.find('p', {'class': 'cast_txt'})
    tomorrowWeather = temp2.text.strip()

    temp2 = temp.find('span', {'class': 'indicator'})
    tomorrowDust = temp2.text.strip().replace('미세먼지 ', '')

    temp = tomorrowBase[1].find('span', {'class': 'todaytemp'})
    tomorrowTemp2 = temp.text.strip()

    temp = tomorrowBase[1].find('ul', {'class': 'info_list'})
    temp2 = temp.find('p', {'class': 'cast_txt'})
    tomorrowWeather2 = temp2.text.strip()

    temp2 = temp.find('span', {'class': 'indicator'})
    tomorrowDust2 = temp2.text.strip().replace('미세먼지 ', '')

    nextWeatherTable = [local, tomorrowTemp, tomorrowWeather, tomorrowDust, tomorrowTemp2, tomorrowWeather2, tomorrowDust2]

    return nextWeatherTable

def setXl():
    file = openpyxl.load_workbook("memoryData.xlsx")
    memoryData = file.active
    for i in range(1, 501):
        memoryData['A' + str(i)].value = '-'
    file.save("memoryData.xlsx")
    file.close()

def setMemory(inputData, outputData):
    returnString = ''
    emptyNum = 0
    file = openpyxl.load_workbook("memoryData.xlsx")
    memoryData = file.active
    for i in range(1, 501):
        if memoryData['A' + str(i)].value == inputData:
            returnString = '{0}는 이미 기억하고 있는걸!'.format(inputData)
            break
        elif memoryData['A' + str(i)].value == '-':
            if memoryData['B' + str(i)].value == None:
                if emptyNum != 0:
                    memoryData['A' + str(emptyNum)].value = inputData
                    memoryData['B' + str(emptyNum)].value = outputData
                    returnString = '{0}는 {1}! 기억! 각인! 세뇌! 삭제!'.format(inputData, outputData)
                    break
                else:
                    memoryData['A' + str(i)].value = inputData
                    memoryData['B' + str(i)].value = outputData
                    returnString = '{0}는 {1}! 기억! 각인! 세뇌! 삭제!'.format(inputData, outputData)
                    break
            else :
                if emptyNum == 0:
                    emptyNum = i

    file.save('memoryData.xlsx')
    file.close()
    return returnString

def deleteMemory(inputData):
    returnString = ''
    file = openpyxl.load_workbook("memoryData.xlsx")
    memoryData = file.active
    for i in range(1, 251):
        if memoryData['A' + str(i)].value == inputData:
            returnString = '{0}...? {1}이었던 거 같은데...으윽 머리가...'.format(memoryData['A' + str(i)].value, memoryData['B' + str(i)].value)
            memoryData['A' + str(i)].value = '-'
            break
        elif memoryData['A' + str(i)].value == '-' and memoryData['B' + str(i)].value == None:
            returnString = '{0}은 기억한 적이 없는데? 늙어서 가르쳐준지도 기억 못하는구나!'.format(inputData)
            break
    file.save('memoryData.xlsx')
    file.close()
    return returnString

def setString(input):
    tmp = [x.strip() for x in input.split('/')]
    return tmp

def findMemory(inputData):
    file = openpyxl.load_workbook("memoryData.xlsx")
    memoryData = file.active
    for i in range(1, 251):
        if memoryData['A' + str(i)].value == inputData:
            tmp = memoryData['B' + str(i)].value
            file.close()
            return tmp
        elif memoryData['A' + str(i)].value == '-' and memoryData['B' + str(i)].value == None:
            file.close()
            return '{0}이 뭔데? 난 모르는데!'.format(inputData)

def printMemory():
    file = openpyxl.load_workbook("memoryData.xlsx")
    memoryData = file.active
    memoryList = []
    for i in range(1, 501):
        if memoryData['A' + str(i)].value != '-' and memoryData['B' + str(i)].value != None:
            memoryList.append('{0} -> {1}'.format(memoryData['A' + str(i)].value, memoryData['B' + str(i)].value))
        elif memoryData['A' + str(i)].value == '-' and memoryData['B' + str(i)].value == None:
            break
    file.close()
    return memoryList

def translate(source, target, text):
    client_id = 'NAVER_CLIENT_ID'
    client_secret = 'NAVER_CLIENT_SECRET'
    encText = urllib.parse.quote(text)
    data = "source={0}&target={1}&text={2}".format(source, target, encText)
    url = "https://openapi.naver.com/v1/papago/n2mt"
    request = urllib.request.Request(url)
    request.add_header("X-Naver-Client-Id",client_id)
    request.add_header("X-Naver-Client-Secret",client_secret)
    response = urllib.request.urlopen(request, data=data.encode("utf-8"))
    rescode = response.getcode()
    if(rescode==200):
        response_body = json.loads(response.read())
        return response_body['message']['result']['translatedText']
    else:
        return "Error Code:" + rescode

def makeTaki():
    file = openpyxl.load_workbook("takiTaki.xlsx", data_only=True)
    takiData = file.active
    tmp = random.randint(1, takiData['B1'].value)
    taki = takiData['A' + str(tmp)].value
    file.close()
    return taki

def makeChaerim():
    file = openpyxl.load_workbook("charims.xlsx", data_only=True)
    charimData = file.active
    tmp = random.randint(1, charimData['B1'].value)
    tmp2 = charimData['C1'].value
    charim = tmp2 + charimData['A' + str(tmp)].value
    file.close()
    return charim

def makeDietTable(dormitory, weekday):
    foodtable = []

    if dormitory == '재정':
        parseDormitory = 'foodtab1'
    elif dormitory == '새롬':
        parseDormitory = 'foodtab2'
    elif dormitory == '이룸':
        parseDormitory = 'foodtab3'

    kangwonURL = 'http://knudorm.kangwon.ac.kr/home/sub02/sub02_05_bj.jsp'
    with urllib.request.urlopen(kangwonURL) as response:
        html = response.read()
        soup = BeautifulSoup(html, 'html.parser')

    dietTable = soup.find('div', {'id': parseDormitory})
    diet1 = dietTable.find_all('table', {'class': 'table_type01'})
    diet2 = diet1[1].find_all('td')

    foodtable.append(diet2[weekday * 3].get_text().replace('\r', '').replace('\t', '').strip())
    foodtable.append(diet2[weekday * 3 + 1].get_text().replace('\r', '').replace('\t', '').strip())
    foodtable.append(diet2[weekday * 3 + 2].get_text().replace('\r', '').replace('\t', '').strip())

    newFoodTable = ['없음' if x == '' else x for x in foodtable]

    return newFoodTable

def makeFortune(name):
    fortuneInt = random.randint(1, 105)
    if fortuneInt > 100:
        return '{0}의 오늘의 행운 지수는 잘 모르겠고 그냥 빛의 가호를 드립니다 ***빛이여!***'.format(name)
    elif fortuneInt > 90:
        return '{0}의 오늘의 행운 지수는 {1}! 복권 하나정도는 사도 될 것 같은 날'.format(name, fortuneInt)
    elif fortuneInt > 80:
        return '{0}의 오늘의 행운 지수는 {1}! 오늘 가챠 연차 대박 각인 날'.format(name, fortuneInt)
    elif fortuneInt > 70:
        return '{0}의 오늘의 행운 지수는 {1}! 길가다가 돈 오백원 정도는 주울 수 있을 거 같은 날'.format(name, fortuneInt)
    elif fortuneInt > 60:
        return '{0}의 오늘의 행운 지수는 {1}! 오늘따라 좋은 템 파밍이 될 거 같은 날'.format(name, fortuneInt)
    elif fortuneInt > 50:
        return '{0}의 오늘의 행운 지수는 {1}! 밖에 나가면 기분이 좋아질 거 같은 날'.format(name, fortuneInt)
    elif fortuneInt > 40:
        return '{0}의 오늘의 행운 지수는 {1}! 좋은것도 아니고 안좋은거도 아닌 그런 평범함으로 채워진거 같은 날'.format(name, fortuneInt)
    elif fortuneInt > 30:
        return '{0}의 오늘의 행운 지수는 {1}! 던전 파밍 해서 허탕만 칠 거 같은 날'.format(name, fortuneInt)
    elif fortuneInt > 20:
        return '{0}의 오늘의 행운 지수는 {1}! 길가다 나한테 D0를 준 교수님 만날 수 있을 거 같은 그런 날'.format(name, fortuneInt)
    elif fortuneInt > 10:
        return '{0}의 오늘의 행운 지수는 {1}! 이어폰 고무 캡 한쪽만 잃어버리지 않게 조심해야 할 거 같은 그런 날'.format(name, fortuneInt)
    else:
        return '{0}의 오늘의 행운 지수는 {1}! 오늘 분명히 뭘 해도 안풀릴거니깐 각오하는게 좋은 날'.format(name, fortuneInt)

def dice(count, value):
    return list(map(lambda x: random.randint(1, value), range(count)))
