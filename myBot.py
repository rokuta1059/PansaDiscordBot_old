#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import sys
import discord
import asyncio
import random
import openpyxl
import json
from discord.ext import commands
from bs4 import BeautifulSoup
import urllib.request
from datetime import datetime, date, time

client = discord.Client()
conch = []
hello = ['왜?', '무슨일?', '하이!', '응!', '짜잔!', '왜 불렀어?', '꺄!', '어엉?', '칫', '뭐']

def makeConch():
    f = open('magicConch.txt', 'r', encoding="utf8")
    readlist = f.read()
    f.close()
    global conch
    conch = readlist.split('\n')

def absurbDiary(inputInt):
    jsonFile = open('diary.json', encoding='utf8').read()
    data = json.loads(jsonFile)
    if inputInt == 0:
        return data[str(random.randint(1, len(data)))]
    else:
        return data[str(inputInt)]

def absurbFindName(name):
    findName = []
    jsonFile = open('diary.json', encoding='utf8').read()
    data = json.loads(jsonFile)
    for i in range(1, len(data) + 1):
        if name in data[str(i)]['name']:
            findName.append(data[str(i)]['number'])
    return findName

def absurbFindAbsurb(search):
    findAb = []
    jsonFile = open('diary.json', encoding='utf8').read()
    data = json.loads(jsonFile)
    for i in range(1, len(data) + 1):
        if search in data[str(i)]['absurb'] or search in data[str(i)]['description']:
            findAb.append(data[str(i)]['number'])
    return findAb

def addAbsurb(absurb, name, descript):
    jsonFile = open('diary.json', encoding='UTF8').read()
    data = json.loads(jsonFile)
    tmp = {'absurb':absurb, 'name':name, 'description':descript, 'number':len(data)+1}
    data[str(len(data)+1)] = tmp

    with open('diary.json', 'w', encoding='UTF8') as saveFile:
        json.dump(data, saveFile, indent='\t', ensure_ascii=False)

    return tmp

def changeAbsurb(absurb, name, descript, num):
    jsonFile = open('diary.json', encoding='UTF8').read()
    data = json.loads(jsonFile)
    tmp = {'absurb':absurb, 'name':name, 'description':descript, 'number':num}
    data[str(num)] = tmp

    with open('diary.json', 'w', encoding='UTF8') as saveFile:
        json.dump(data, saveFile, indent='\t', ensure_ascii=False)

    return tmp

def cypersRank(nickname, game):
    parseNick = urllib.parse.quote(nickname)
    if game == "공식":
        cypersURL = 'http://cyphers.nexon.com/cyphers/game/record/search/1/' + parseNick + '/1'
    else:
        cypersURL = 'http://cyphers.nexon.com/cyphers/game/record/search/2/' + parseNick + '/2'
    with urllib.request.urlopen(cypersURL) as response:
        html = response.read()
        soup = BeautifulSoup(html, 'html.parser')

    a = []
    battle = []
    img = []

    for firstTable in soup.find_all("div", "info info1"):
        for firstValue in firstTable.find_all("td"):
            a.append(firstValue.get_text())

    for secondTable in soup.find_all("div", "info info2"):
        for secondValue in secondTable.find_all("td"):
            a.append(secondValue.get_text())

    for battleTable in soup.find_all("li", "show"):
        for battleResult in battleTable.find_all("td"):
            battle.append(battleResult.get_text().replace(
                '\n', '').replace('\r', '').replace('\t', ''))

        for battleResult in battleTable.find_all("td", "char"):
            for imgList in battleResult.find_all("img"):
                img.append(imgList.get("src"))

    return cypersURL, a, battle, img

def getMillDate(name):
    f = open('militaryDate.txt', 'r', encoding="utf8")
    mill = []
    findName = False
    while True:
        line = f.readline()
        if line.find(name) != -1:
            mill = line.replace('\n', '').split(' ')
            findName = True
            break
        if not line:
            break
    f.close()

    if findName:
        milldate = list(map(int, mill[1].split('.')))
        now = datetime.now()
        findDate = datetime(milldate[0], milldate[1], milldate[2]) - now
        return "{0}쿤의 전역일은 **{1}일** 남았답니다~".format(mill[0], findDate.days)
    else:
        return "{0}쿤은 군인이 아닌거 같아요~".format(name)

def todayWeather(search):
    client_id = 'NAVER_CLIENT_ID'
    client_secret = 'NAVER_CLIENT_SECRET'
    encText = urllib.parse.quote(search + ' 날씨')
    url = 'https://search.naver.com/search.naver?sm=top_hty&fbm=1&ie=utf8&query=' + encText
    request = urllib.request.Request(url)
    request.add_header("X-Naver-Client-Id",client_id)
    request.add_header("X-Naver-Client-Secret",client_secret)
    response = urllib.request.urlopen(request)
    bsObj = BeautifulSoup(response, "html.parser")
    todayBase = bsObj.find('div', {'class': 'today_area _mainTabContent'})

    temp = bsObj.find('div', {'class': 'select_box'})
    temp2 = temp.find('em')
    local = temp2.text.strip()

    temp = todayBase.find('span', {'class': 'todaytemp'})
    todayTemp = temp.text.strip()

    temp = todayBase.find('p', {'class': 'cast_txt'})
    todayCast = temp.text.strip()

    temp = todayBase.find('span', {'class': 'merge'})
    todayMerge = temp.text.strip()

    temp = todayBase.find('span', {'class': 'sensible'})
    temp2 = temp.find('span', {'class': 'num'})
    todaySensible = temp2.text.strip()

    temp = todayBase.find('dl', {'class': 'indicator'})
    temp2 = temp.find_all('dd')
    todayDust = temp2[0].text.strip()

    weatherTable = [local, todayTemp, todayCast, todayMerge, todaySensible, todayDust]

    return weatherTable

def tomorrowWeather(search):
    client_id = 'NAVER_CLIENT_ID'
    client_secret = 'NAVER_CLIENT_SECRET'
    encText = urllib.parse.quote(search + ' 날씨')
    url = 'https://search.naver.com/search.naver?sm=top_hty&fbm=1&ie=utf8&query=' + encText
    request = urllib.request.Request(url)
    request.add_header("X-Naver-Client-Id",client_id)
    request.add_header("X-Naver-Client-Secret",client_secret)
    response = urllib.request.urlopen(request)
    bsObj = BeautifulSoup(response, "html.parser")

    temp = bsObj.find('div', {'class': 'select_box'})
    temp2 = temp.find('em')
    local = temp2.text.strip()

    tomorrowBase = bsObj.find_all('div', {'class': 'main_info morning_box'})

    temp = tomorrowBase[0].find('span', {'class': 'todaytemp'})
    tomorrowTemp = temp.text.strip()

    temp = tomorrowBase[0].find('ul', {'class': 'info_list'})
    temp2 = temp.find('p', {'class': 'cast_txt'})
    tomorrowWeather = temp2.text.strip()

    temp2 = temp.find('span', {'class': 'indicator'})
    tomorrowDust = temp2.text.strip().replace('미세먼지 ', '')

    temp = tomorrowBase[1].find('span', {'class': 'todaytemp'})
    tomorrowTemp2 = temp.text.strip()

    temp = tomorrowBase[1].find('ul', {'class': 'info_list'})
    temp2 = temp.find('p', {'class': 'cast_txt'})
    tomorrowWeather2 = temp2.text.strip()

    temp2 = temp.find('span', {'class': 'indicator'})
    tomorrowDust2 = temp2.text.strip().replace('미세먼지 ', '')

    nextWeatherTable = [local, tomorrowTemp, tomorrowWeather, tomorrowDust, tomorrowTemp2, tomorrowWeather2, tomorrowDust2]

    return nextWeatherTable

def setXl():
    file = openpyxl.load_workbook("memoryData.xlsx")
    memoryData = file.active
    for i in range(1, 501):
        memoryData['A' + str(i)].value = '-'
    file.save("memoryData.xlsx")
    file.close()

def setMemory(inputData, outputData):
    returnString = ''
    emptyNum = 0
    file = openpyxl.load_workbook("memoryData.xlsx")
    memoryData = file.active
    for i in range(1, 501):
        if memoryData['A' + str(i)].value == inputData:
            returnString = '{0}는 이미 기억하고 있는걸!'.format(inputData)
            break
        elif memoryData['A' + str(i)].value == '-':
            if memoryData['B' + str(i)].value == None:
                if emptyNum != 0:
                    memoryData['A' + str(emptyNum)].value = inputData
                    memoryData['B' + str(emptyNum)].value = outputData
                    returnString = '{0}는 {1}! 기억! 각인! 세뇌! 삭제!'.format(inputData, outputData)
                    break
                else:
                    memoryData['A' + str(i)].value = inputData
                    memoryData['B' + str(i)].value = outputData
                    returnString = '{0}는 {1}! 기억! 각인! 세뇌! 삭제!'.format(inputData, outputData)
                    break
            else :
                if emptyNum == 0:
                    emptyNum = i

    file.save('memoryData.xlsx')
    file.close()
    return returnString

def deleteMemory(inputData):
    returnString = ''
    file = openpyxl.load_workbook("memoryData.xlsx")
    memoryData = file.active
    for i in range(1, 251):
        if memoryData['A' + str(i)].value == inputData:
            returnString = '{0}...? {1}이었던 거 같은데...으윽 머리가...'.format(memoryData['A' + str(i)].value, memoryData['B' + str(i)].value)
            memoryData['A' + str(i)].value = '-'
            break
        elif memoryData['A' + str(i)].value == '-' and memoryData['B' + str(i)].value == None:
            returnString = '{0}은 기억한 적이 없는데? 늙어서 가르쳐준지도 기억 못하는구나!'.format(inputData)
            break
    file.save('memoryData.xlsx')
    file.close()
    return returnString

def setString(input):
    tmp = [x.strip() for x in input.split('/')]
    return tmp

def findMemory(inputData):
    file = openpyxl.load_workbook("memoryData.xlsx")
    memoryData = file.active
    for i in range(1, 251):
        if memoryData['A' + str(i)].value == inputData:
            tmp = memoryData['B' + str(i)].value
            file.close()
            return tmp
        elif memoryData['A' + str(i)].value == '-' and memoryData['B' + str(i)].value == None:
            file.close()
            return '{0}이 뭔데? 난 모르는데!'.format(inputData)

def printMemory():
    file = openpyxl.load_workbook("memoryData.xlsx")
    memoryData = file.active
    memoryList = []
    for i in range(1, 501):
        if memoryData['A' + str(i)].value != '-' and memoryData['B' + str(i)].value != None:
            memoryList.append('{0} -> {1}'.format(memoryData['A' + str(i)].value, memoryData['B' + str(i)].value))
        elif memoryData['A' + str(i)].value == '-' and memoryData['B' + str(i)].value == None:
            break
    file.close()
    return memoryList

def translate(source, target, text):
    client_id = 'NAVER_CLIENT_ID'
    client_secret = 'NAVER_CLIENT_SECRET'
    encText = urllib.parse.quote(text)
    data = "source={0}&target={1}&text={2}".format(source, target, encText)
    url = "https://openapi.naver.com/v1/papago/n2mt"
    request = urllib.request.Request(url)
    request.add_header("X-Naver-Client-Id",client_id)
    request.add_header("X-Naver-Client-Secret",client_secret)
    response = urllib.request.urlopen(request, data=data.encode("utf-8"))
    rescode = response.getcode()
    if(rescode==200):
        response_body = json.loads(response.read())
        return response_body['message']['result']['translatedText']
    else:
        return "Error Code:" + rescode

def makeTaki():
    file = openpyxl.load_workbook("takiTaki.xlsx", data_only=True)
    takiData = file.active
    tmp = random.randint(1, takiData['B1'].value)
    taki = takiData['A' + str(tmp)].value
    file.close()
    return taki

def makeChaerim():
    file = openpyxl.load_workbook("charims.xlsx", data_only=True)
    charimData = file.active
    tmp = random.randint(1, charimData['B1'].value)
    tmp2 = charimData['C1'].value
    charim = tmp2 + charimData['A' + str(tmp)].value
    file.close()
    return charim

def makeDietTable(dormitory, weekday):
    foodtable = []

    if dormitory == '재정':
        parseDormitory = 'foodtab1'
    elif dormitory == '새롬':
        parseDormitory = 'foodtab2'
    elif dormitory == '이룸':
        parseDormitory = 'foodtab3'

    kangwonURL = 'http://knudorm.kangwon.ac.kr/home/sub02/sub02_05_bj.jsp'
    with urllib.request.urlopen(kangwonURL) as response:
        html = response.read()
        soup = BeautifulSoup(html, 'html.parser')

    dietTable = soup.find('div', {'id': parseDormitory})
    diet1 = dietTable.find_all('table', {'class': 'table_type01'})
    diet2 = diet1[1].find_all('td')

    foodtable.append(diet2[weekday * 3].get_text().replace('\r', '').replace('\t', '').strip())
    foodtable.append(diet2[weekday * 3 + 1].get_text().replace('\r', '').replace('\t', '').strip())
    foodtable.append(diet2[weekday * 3 + 2].get_text().replace('\r', '').replace('\t', '').strip())

    newFoodTable = ['없음' if x == '' else x for x in foodtable]

    return newFoodTable

def makeFortune(name):
    fortuneInt = random.randint(1, 105)
    if fortuneInt > 100:
        return '{0}의 오늘의 행운 지수는 잘 모르겠고 그냥 빛의 가호를 드립니다 ***빛이여!***'.format(name)
    elif fortuneInt > 90:
        return '{0}의 오늘의 행운 지수는 {1}! 복권 하나정도는 사도 될 것 같은 날'.format(name, fortuneInt)
    elif fortuneInt > 80:
        return '{0}의 오늘의 행운 지수는 {1}! 오늘 가챠 연차 대박 각인 날'.format(name, fortuneInt)
    elif fortuneInt > 70:
        return '{0}의 오늘의 행운 지수는 {1}! 길가다가 돈 오백원 정도는 주울 수 있을 거 같은 날'.format(name, fortuneInt)
    elif fortuneInt > 60:
        return '{0}의 오늘의 행운 지수는 {1}! 오늘따라 좋은 템 파밍이 될 거 같은 날'.format(name, fortuneInt)
    elif fortuneInt > 50:
        return '{0}의 오늘의 행운 지수는 {1}! 밖에 나가면 기분이 좋아질 거 같은 날'.format(name, fortuneInt)
    elif fortuneInt > 40:
        return '{0}의 오늘의 행운 지수는 {1}! 좋은것도 아니고 안좋은거도 아닌 그런 평범함으로 채워진거 같은 날'.format(name, fortuneInt)
    elif fortuneInt > 30:
        return '{0}의 오늘의 행운 지수는 {1}! 던전 파밍 해서 허탕만 칠 거 같은 날'.format(name, fortuneInt)
    elif fortuneInt > 20:
        return '{0}의 오늘의 행운 지수는 {1}! 길가다 나한테 D0를 준 교수님 만날 수 있을 거 같은 그런 날'.format(name, fortuneInt)
    elif fortuneInt > 10:
        return '{0}의 오늘의 행운 지수는 {1}! 이어폰 고무 캡 한쪽만 잃어버리지 않게 조심해야 할 거 같은 그런 날'.format(name, fortuneInt)
    else:
        return '{0}의 오늘의 행운 지수는 {1}! 오늘 분명히 뭘 해도 안풀릴거니깐 각오하는게 좋은 날'.format(name, fortuneInt)

def dice(count, value):
    return list(map(lambda x: random.randint(1, value), range(count)))

@client.event
async def on_ready():
    print('이몸 등장이올시다')
    print(client.user.name)
    print(client.user.id)
    print('------')
    await client.change_presence(activity=discord.Game(name='동아리방 먼지와 함께 뒹굴뒹굴', type=1))

@client.event
async def on_message(message):

    if message.content.startswith('!도움'):
        embed = discord.Embed(
            title='**판 사 요 정 등 장**',
            description='이거슨 판사 요정 커맨드 리스트',
            colour=discord.Colour.green()
        )
        embed.set_thumbnail(url=client.user.avatar_url)
        embed.add_field(name='!망언, !diary', value='판사 역대 망언 출력해줌. 숫자 같이 입력하면 해당 번호 망언 나옴.')
        embed.add_field(name='!망언검색 (이름/내용) (검색어)', value='망언 검색해줌. 몇번인지 나옴.')
        embed.add_field(name='!선택', value='선택 장애인들이 넘치는 동아리를 위한 멋진 커맨드. 띄어쓰기로 나눠서 이것저것 입력하면 하나 골라줌.')
        embed.add_field(name='!타키, !taki', value='오늘의 타키쿤은?')
        embed.add_field(name='!채림, !채릠, !chaerim', value='오늘의 채림쓰는?')
        embed.add_field(name='!소라고둥, 소라고둥님, 마법의소라고둥님, 마법의 소라고둥님', value='위대하신 소라고둥님의 말을 들을 수 있다.')
        embed.add_field(name='!전역일', value='군인님들 본명 OOO 입력하면 몇일 남았는지 알려줌. 추가되고 싶은 사람은 쥔장한테 ㄱㄱ')
        embed.add_field(name='!날씨, !오늘날씨', value='오늘날씨 알려줌. 지역 같이 입력하면 됨.')
        embed.add_field(name='!내일날씨', value='내일날씨 알려줌. 지역 같이 입력하면 됨.')
        embed.add_field(name='!사이퍼즈 (닉네임)', value='전적 알려줌.')
        embed.add_field(name='!사이퍼즈 공식 (닉네임)', value='전적이랑 공식전 최근에 한거 결과 알려줌.')
        embed.add_field(name='!사이퍼즈 일반 (닉네임)', value='전적이랑 일반전 최근에 한거 결과 알려줌.')
        embed.add_field(name='!기억', value='/로 구분해서 기억함. A/B 입력하면 A 에 B가 대응')
        embed.add_field(name='!삭제', value='기억 삭제함. A 입력하면 A에 대응하는 걸 잊음')
        embed.add_field(name='!말해 or 요정아', value='기억한거 말함. A 입력하면 대응하는 B 출력')
        embed.add_field(name='!리스트', value='기억 리스트 나옴. A->B')
        embed.add_field(name='!한영, !영한, !한일, !일한', value='번역함. 네이버 파파고 제공')
        embed.add_field(name='!재정, !새롬, !이룸', value='오늘 밥 뭔지 알려줌.')
        embed.add_field(name='!운세, !fortune', value='아주 간단한 오늘의 운세')
        embed.add_field(name='!요정수치, !fairy', value='그때그때 요정수치')
        embed.add_field(name='!주사위, !dice, ㅈㅅㅇ (굴릴 횟수)D(몇면체) + ...', value='주사위 굴림. 총합도 알려줌.')
        embed.add_field(name='!모두모여', value='뒷문장 추가해서 전체멘션')
        embed.set_footer(text='강원대 판화사랑 동아리 컴정 15학번 과잠선배 제작')
        await message.channel.send(embed=embed)

    if message.content.startswith('!!공지'):
        mes = message.content.replace('!!공지', '').strip().split("/")
        member = message.author
        embed = discord.Embed(
            title='{}'.format(mes[0]),
            description='{}'.format(mes[1]),
            colour=discord.Colour.blurple()
        )
        embed.set_footer(text='{0}이(가) 모두에게 전달중'.format(member.name),
            icon_url=member.avatar_url)
        await client.get_channel('TEXT_CHANNEL_ID').send(embed=embed)

    if message.content.startswith('!망언') or message.content.startswith('!diary'):
        if message.content.startswith('!망언검색'):
            mes = message.content.split(" ")
            if mes[1] in '이름':
                numList = absurbFindName(mes[2])
                numStr = ''
                for i in numList:
                    numStr = numStr + str(i) + ' '
                embed = discord.Embed(
                    title='{0}이(가) 이름에 포함된 망언은...'.format(mes[2]),
                    description=numStr,
                    colour=random.randint(0, 0xffffff)
                )
                await message.channel.send(embed=embed)
            elif mes[1] in '내용':
                numList = absurbFindAbsurb(mes[2])
                numStr = ''
                for i in numList:
                    numStr = numStr + str(i) + ' '
                embed = discord.Embed(
                    title='{0}이(가) 내용에 포함된 망언은...'.format(mes[2]),
                    description=numStr,
                    colour=random.randint(0, 0xffffff)
                )
                await message.channel.send(embed=embed)
        elif message.content.startswith('!망언추가'):
            mes = message.content.replace('!망언추가', '').strip().split('/')
            absurb = addAbsurb(mes[0], mes[1], mes[2])
            embed = discord.Embed(
                title='{0}. {1}'.format(absurb['number'], absurb['absurb']),
                description='- ***{0}***, *{1}*'.format(absurb['name'], absurb['description']),
                colour=random.randint(0, 0xffffff)
            )
            await message.channel.send(embed=embed)
        elif message.content.startswith('!망언수정'):
            mes = message.content.replace('!망언수정', '').strip().split('/')
            absurb = changeAbsurb(mes[1], mes[2], mes[3], int(mes[0]))
            embed = discord.Embed(
                title='{0}. {1}'.format(absurb['number'], absurb['absurb']),
                description='- ***{0}***, *{1}*'.format(absurb['name'], absurb['description']),
                colour=random.randint(0, 0xffffff)
            )
            await message.channel.send(embed=embed)
        else:
            mes = message.content.split(" ")
            if len(mes) == 2:
                absurb = absurbDiary(int(mes[1]))
            else:
                absurb = absurbDiary(0)
            embed = discord.Embed(
                title='{0}. {1}'.format(absurb['number'], absurb['absurb']),
                description='- ***{0}***, *{1}*'.format(absurb['name'], absurb['description']),
                colour=random.randint(0, 0xffffff)
            )
            await message.channel.send(embed=embed)
    
    if message.content.startswith('!타키') or message.content.startswith('!taki'):
        await message.channel.send(makeTaki())

    if message.content.startswith('!채림') or message.content.startswith('!채릠') or message.content.startswith('!chaerim'):
        await message.channel.send(makeChaerim())

    if message.content.startswith('!선택'):
        mes = message.content.split(" ")
        del mes[0]
        await message.channel.send(random.choice(mes))

    if message.content.startswith('!소라고둥') or message.content.startswith('소라고둥님') or message.content.startswith('마법의소라고둥님') or message.content.startswith('마법의 소라고둥님'):
        await message.channel.send(random.choice(conch))

    if message.content.startswith('!전역일'):
        mes = message.content.split(" ")
        await message.channel.send(getMillDate(mes[1]))

    if message.content.startswith('!사이퍼즈'):
        mes = message.content.split(" ")
        if mes[1] == '일반' or mes[1] == '공식':
            nickname = mes[2]
            nickURL, log, battleResult, imgURL = cypersRank(nickname, mes[1])
        else:
            nickname = mes[1]
            nickURL, log, battleResult, imgURL = cypersRank(nickname, '')

        embed = discord.Embed(
            title='**CYPHERS**',
            description='사이퍼즈 전적입니다',
            colour=discord.Colour.red(),
        )
        if len(log) <= 4:
            embed.set_thumbnail(
                url='http://static.cyphers.co.kr/img/event/logo_bar.gif')
            embed.add_field(name='닉네임', value=log[0])
            embed.add_field(name='급수', value=log[1])
            embed.add_field(name='클랜', value=log[2])
            embed.add_field(name='승패 RP 등등', value='이번 시즌 공식전 기록이 없는 거시와요')
            await message.channel.send(embed=embed)

        else:
            embed.set_thumbnail(
                url='http://static.cyphers.co.kr/img/event/logo_bar.gif')
            embed.add_field(name='닉네임', value=log[0])
            embed.add_field(name='급수', value=log[1])
            embed.add_field(name='클랜', value=log[2])
            embed.add_field(name='승패', value=log[3])
            embed.add_field(name='공식전 RP', value=log[4])
            embed.add_field(name='최고 RP', value=log[5])
            embed.add_field(name='티어', value=log[6])
            await message.channel.send(embed=embed)

        if mes[1] == '일반' or mes[1] == '공식':
            embed = discord.Embed(
                title='**가장 최근 경기 결과**',
                url=nickURL,
                description='가장 최근에 플레이한 경기의 결과입니다.',
                colour=discord.Colour.red()
            )
            embed.set_thumbnail(url=imgURL[0])
            embed.add_field(name='플레이 시간 및 결과', value=battleResult[0])
            embed.add_field(name='캐릭터', value=battleResult[1])
            embed.add_field(name='레벨', value=battleResult[2])
            embed.add_field(name='킬', value=battleResult[3])
            embed.add_field(name='데스', value=battleResult[4])
            embed.add_field(name='도움', value=battleResult[5])
            embed.add_field(name='공격량', value=battleResult[6])
            embed.add_field(name='피해량', value=battleResult[7])
            embed.add_field(name='전투참여', value=battleResult[8])
            embed.add_field(name='시야확보', value=battleResult[9])
            await message.channel.send(embed=embed)

    if message.content.startswith('!오늘날씨') or message.content.startswith('!날씨'):
        mes = message.content.replace('!오늘날씨', '').replace('!날씨', '').strip()
        today = todayWeather(mes)
        embed = discord.Embed(
                title='**날씨! 오늘! 날씨!**',
                description=today[0]+'의 오늘 날씨!',
                colour=discord.Colour.blue()
            )
        embed.add_field(name='현재 기온', value=today[1]+'도')
        embed.add_field(name='상태', value=today[2])
        embed.add_field(name='최저최고기온', value=today[3])
        embed.add_field(name='체감온도', value=today[4])
        embed.add_field(name='미세먼지', value=today[5])
        embed.set_footer(text='네이버 날씨 제공')
        await message.channel.send(embed=embed)

    if message.content.startswith('!내일날씨'):
        mes = message.content.replace('!내일날씨', '').strip()
        tomorr = tomorrowWeather(mes)
        embed = discord.Embed(
                title='**날씨! 내일! 날씨!**',
                description=tomorr[0]+'의 내일 날씨!',
                colour=discord.Colour.blue()
            )
        embed.add_field(name='내일 오전', value=tomorr[1]+'도')
        embed.add_field(name='상태', value=tomorr[2])
        embed.add_field(name='미세먼지', value=tomorr[3])
        embed.add_field(name='내일 오후', value=tomorr[4]+'도')
        embed.add_field(name='상태', value=tomorr[5])
        embed.add_field(name='미세먼지', value=tomorr[6])
        embed.set_footer(text='네이버 날씨 제공')
        await message.channel.send(embed=embed)
    
    if message.content.startswith('!기억'):
        mes = message.content.replace('!기억', '')
        mesTemp = setString(mes)
        if mesTemp[0] == '' or mesTemp[1] == '':
            await message.channel.send('기억할 문장이 이상한 거 같은데...제대로 가르쳐줘!')
        else:
            result = setMemory(mesTemp[0], mesTemp[1])
            await message.channel.send(result)
    
    if message.content.startswith('!삭제'):
        mes = message.content.replace('!삭제', '').strip()
        result = deleteMemory(mes)
        await message.channel.send(result)
    
    if message.content.startswith('!말해'):
        mes = message.content.replace('!말해', '').strip()
        result = findMemory(mes)
        await message.channel.send(result)

    if message.content.startswith('요정아'):
        mes = message.content.replace('요정아', '').strip()
        if mes == '':
            await message.channel.send(random.choice(hello))
        else :
            result = findMemory(mes)
            await message.channel.send(result)
    
    if message.content.startswith('!리스트'):
        memoryDataList = printMemory()
        printListStr = "```"
        for i in memoryDataList:
            printListStr = printListStr + i + '\n'
        printListStr = printListStr + "```"
        await message.channel.send('내 머리속을 보고 싶어하다니...변태...')
        await message.channel.send(printListStr)

    if message.content.startswith('!한일'):
        mes = message.content.replace('!한일', '').strip()
        transText = translate('ko', 'ja', mes)
        embed = discord.Embed(
            title=transText,
            description=mes,
            colour=discord.Colour.purple()
        )
        embed.set_footer(text='Translated by.네이버 파파고')
        await message.channel.send(embed=embed)
    
    if message.content.startswith('!일한'):
        mes = message.content.replace('!일한', '').strip()
        transText = translate('ja', 'ko', mes)
        embed = discord.Embed(
            title=transText,
            description=mes,
            colour=discord.Colour.purple()
        )   
        embed.set_footer(text='Translated by.네이버 파파고')
        await message.channel.send(embed=embed)

    if message.content.startswith('!한영'):
        mes = message.content.replace('!한영', '').strip()
        transText = translate('ko', 'en', mes)
        embed = discord.Embed(
            title=transText,
            description=mes,
            colour=discord.Colour.purple()
        )   
        embed.set_footer(text='Translated by.네이버 파파고')
        await message.channel.send(embed=embed)

    if message.content.startswith('!영한'):
        mes = message.content.replace('!영한', '').strip()
        transText = translate('en', 'ko', mes)
        embed = discord.Embed(
            title=transText,
            description=mes,
            colour=discord.Colour.purple()
        )   
        embed.set_footer(text='Translated by.네이버 파파고')
        await message.channel.send(embed=embed)

    if message.content.startswith('!재정') or message.content.startswith('!새롬') or message.content.startswith('!이룸'):
        dietTable = makeDietTable(message.content.replace('!', '').strip(), datetime.today().weekday())
        embed = discord.Embed(
            title = '식 단 표',
            description = '{0}의 오늘의 식단표'.format(message.content.replace('!', '').strip()),
            colour=discord.Colour.blue()
        )
        embed.add_field(name='아침', value = dietTable[0])
        embed.add_field(name='점심', value = dietTable[1])
        embed.add_field(name='저녁', value = dietTable[2])
        await message.channel.send(embed=embed)
    
    if message.content.startswith('!모두모여'):
        mes = message.content.replace('!모두모여', '').strip()
        await message.channel.send('@everyone ' + mes)

    if message.content.startswith('!운세') or message.content.startswith('!fortune'):
        await message.channel.send(makeFortune(message.author.display_name))

    if message.content.startswith('!fairy') or message.content.startswith('!요정수치'):
        await message.channel.send('{0}의 동방 요정수치는 무려 >>>{1}<<<씩이나 된다구?'.format(message.author.display_name, random.randint(1, 9999)))

    if message.content.startswith('!dice') or message.content.startswith('!주사위') or message.content.startswith("ㅈㅅㅇ"):
        mes = message.content.replace("!dice", "").replace("!주사위", "").replace("ㅈㅅㅇ", "").strip().upper().split('+')
        result = 0
        embed = discord.Embed(
            title = '주사위 결과',
            description = '입력한 주사위의 결과와 합입니다',
            color = random.randint(0, 0xffffff)
        )
        for i in mes:
            tmpDice = i.split('D')
            diceResult = dice(int(tmpDice[0]), int(tmpDice[1]))
            embed.add_field(name='{0}주사위'.format(i), value=diceResult)
            embed.add_field(name='{0}주사위의 합'.format(i), value=sum(diceResult))
            result = result + sum(diceResult)
        embed.add_field(name='주사위 총 합', value='{}'.format(result))
        await message.channel.send(embed=embed)

makeConch()
client.run('TOKEN')
